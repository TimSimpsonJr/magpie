"""Robust dirty-data loader for FOIA CSV / XLSX exports.

FOIA and audit-log exports arrive dirty. A county system writes latin-1, an
upstream Excel round-trip strips the leading zero off every ZIP, empty cells
that mean "missing" come back as ``""``, and a spreadsheet encodes a category
label as a vertical merge that the reader only stores in the top-left cell.
:func:`load_table` turns one such file into a clean :class:`pandas.DataFrame`
plus a structured :class:`LoadResult.report` describing what it had to do.

Design notes (verified against the Phase 3 research gate,
``skills/dataset-analyze/references/prior-art.md``, on the pinned dev venv):

* **Encoding pre-flight** uses ``charset-normalizer`` (already a transitive dep
  of ``requests``) to sniff a byte sample when the caller does not pin an
  encoding. Single-byte / legacy codepages are INHERENTLY AMBIGUOUS: a short
  sample can score several of them (``cp775``, ``latin-1``, ``cp1250``, ...)
  identically because they decode the same bytes into *different* characters
  without raising -- so there is no reliable silent auto-pick, and the detector
  reporting high confidence does not mean it picked the right one. Detection is
  therefore HONEST rather than silently confident: the report flags ambiguity
  (``encoding_low_confidence``) and lists the rival candidates
  (``encoding_alternatives``) so a caller/skill can require an explicit
  ``encoding=``. Pin ``encoding=`` (e.g. ``"latin-1"``) for a byte-exact read;
  the best-effort decode is still returned, but never presented as certain.
* **TEXT-whitelist.** Columns whose name is ID-like load as strings so leading
  zeros survive (``07054`` stays ``"07054"``, not the int ``7054``). Matching is
  on TOKEN BOUNDARIES, not raw substrings: the name is lowercased and split on
  separators and camelCase humps, then a pattern matches only when it equals a
  WHOLE token. This single rule applies UNIFORMLY to every pattern (``id``,
  ``case``, ``plate``, ``ssn``, ``dob``, ``zip``, ``zipcode``, ``fips``,
  ``phone``, ``account``, ``license``), so it stops real numerics like ``valid``,
  ``incident_count``, ``candidate_total``, ``casein`` AND words that merely
  contain a longer pattern -- ``microphone_level`` (phone), ``zipper_count``
  (zip), ``accountability_score`` (account) -- from being coerced to text (which
  would break the ``stats.py`` consumer), while still catching ``zip_code``,
  ``zipcode``, ``phone_number``, ``account_id``, ``fips``, ``caseNumber``,
  ``Account_ID``. Any name the caller passes in ``text_columns`` is always
  honored (the escape hatch for an edge-case ID column the token rule misses),
  and every coerced column is listed in ``report["text_columns"]`` so a wrong
  one is visible. On the CSV path the whitelist is ``dtype=str`` for those
  columns; on the XLSX path it is post-read coercion (pandas/openpyxl would
  otherwise int-ify them).
* **``empty_null``.** When True (default), ONLY a literal empty / whitespace
  cell reads as missing (``NaN``), consistently across both paths. The NA set is
  deliberately NARROW: a cell that literally contains ``N/A`` / ``NULL`` /
  ``NaN`` / ``None`` survives as that STRING, because pandas' default-NA
  spellings are NOT applied (the CSV read pins ``keep_default_na=False``). This
  keeps the downstream ``***``-vs-blank-vs-NULL rigor distinction honest (a
  redaction is present; a blank is absent; a literal ``NULL`` token is its own
  signal, not silently erased). When False, the caller opts out of ``""``->NA
  entirely: empty cells are kept literally as ``""`` (so a column that would
  otherwise have been numeric is read as object to hold the ``""``), and the
  default-NA spellings are likewise preserved as literal strings.
* **XLSX merged cells.** Merged regions are *unmerged then filled* from the
  top-left value (openpyxl drops every non-top-left cell on merge), so a
  vertically-merged label populates every row of its region. pandas 3.0.3
  removed ``fillna(method="ffill")`` (it raises ``TypeError``); the fallback
  forward-fill uses ``df.ffill()``.
* **Junk headers.** ``skiprows`` (int or list of 0-based indices) drops
  preamble rows before the real header.
* **Parquet cache.** When ``parquet_cache`` is given, the cleaned frame is
  written to Parquet (via pandas/pyarrow) and the path is noted in the report;
  a follow-up call could read it back.

The load itself uses ``pd.read_csv`` with per-column ``dtype=str`` (correct and
sufficient: it reads ID columns as text bytes losslessly, no double-rounding).
DuckDB is intentionally NOT used here -- it is reserved for the Phase 3.3/3.4
columnar cache / query layer, downstream of this pure loader.

The loader is pure except for file IO. It is decoupled from ``stats.py`` and
from any corpus-specific loader: its only inputs are a path and options.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

# Column-NAME matching for ID-like fields whose leading zeros must be preserved.
# Matching is on TOKEN BOUNDARIES, never raw substrings: the name is lowercased
# and split on non-alphanumeric separators AND camelCase / letter-digit
# boundaries (so "Account_ID" -> [account, id], "caseNumber" -> [case, number],
# "zip_code" -> [zip, code]). A pattern matches ONLY when it equals a WHOLE
# token -- this rule is applied UNIFORMLY to every pattern, short or long.
#
# Why uniform whole-token (not substring for the "unambiguous" long ones): a
# substring test mis-fires on perfectly ordinary words that happen to CONTAIN a
# pattern -- "microphone_level" contains "phone", "zipper_count" contains "zip",
# "accountability_score" contains "account". Coercing those real (often numeric)
# columns to text silently breaks the stats.py consumer, which casts the column
# str -> float (gini/rates). Whole-token matching rejects all three while still
# catching "zip"/"zip_code"/"zipcode"/"ZipCode", "phone_number", "account_id",
# "fips", "case"/"case_number"/"caseNumber", "ssn", "plate", "dob".
#
# Single-token spellings are listed explicitly so a one-word column name still
# matches: "zipcode" (no separator to split) is its own entry alongside "zip".
#
# Edge cases an operator can steer:
#  * A column whose name genuinely contains an ID token but is NOT an ID (rare,
#    e.g. a hypothetical "license_plateau") would be coerced to text. That is the
#    HARMLESS direction (a string of characters), and `.report["text_columns"]`
#    lists every coerced column so a wrong one is visible. To force such a column
#    numeric, the caller leaves it out of `text_columns` (auto-coercion of a true
#    token is intentional); to force an unmatched edge-case ID column TO text,
#    pass it in `text_columns`.
#
# A false positive only forces a column to text (harmless: a string of digits);
# a false negative silently drops a leading zero (the failure we guard against).
ID_LIKE_TOKEN_PATTERNS: frozenset[str] = frozenset(
    {
        "id",
        "case",
        "plate",
        "ssn",
        "dob",
        "zip",
        "zipcode",
        "fips",
        "phone",
        "account",
        "license",
    }
)

# Split on runs of non-alphanumerics, camelCase humps (lower/digit -> upper),
# acronym tails (UPPER before Upper+lower, e.g. "SSNField" -> SSN|Field), and
# letter<->digit boundaries ("zip5" -> zip|5). Implemented by inserting a marker
# at each boundary, then splitting on the marker plus separators.
_CAMEL_BOUNDARY_RE = re.compile(
    r"(?<=[a-z0-9])(?=[A-Z])"          # fooBar  -> foo|Bar
    r"|(?<=[A-Z])(?=[A-Z][a-z])"        # SSNFile -> SSN|File
    r"|(?<=[A-Za-z])(?=[0-9])"          # zip5    -> zip|5
    r"|(?<=[0-9])(?=[A-Za-z])"          # 5zip    -> 5|zip
)
_NON_ALNUM_RE = re.compile(r"[^0-9a-z]+")

# Empty-string spellings that ``empty_null`` collapses to missing. Kept narrow
# on purpose: only a literal empty cell (and surrounding whitespace) is treated
# as absent. Sentinel tokens like "***", "N/A", "NULL" are NOT swept here --
# distinguishing a redaction ("***", a value that exists but is withheld) from a
# genuine blank is a downstream rigor concern, and over-eager NA coercion would
# erase that signal. This narrowness is ENFORCED on the CSV read by pinning
# keep_default_na=False (so pandas does not add its default N/A / NULL / NaN /
# None spellings on top of this set); see _load_csv.
_EMPTY_NA_VALUES: tuple[str, ...] = ("", " ")


@dataclass
class LoadResult:
    """The cleaned table plus a structured report of how it was loaded.

    Attributes:
        df: the cleaned :class:`pandas.DataFrame`.
        report: a dict describing the load. Keys:

            * ``encoding`` -- the encoding actually used to read the file
              (``"binary"`` for XLSX).
            * ``encoding_used`` -- alias of ``encoding`` (the encoding the read
              actually used), exposed under an explicit name for callers.
            * ``encoding_detected`` -- ``True`` if it was sniffed,
              ``False`` if the caller pinned it (or it was a native binary
              read, as for XLSX).
            * ``encoding_confidence`` -- detector confidence in ``[0, 1]`` when
              sniffed, else ``None``. NOTE: this does NOT discriminate among
              single-byte alternatives (cp775 and latin-1 can both score
              ``1.0`` on the same bytes); always read it together with
              ``encoding_low_confidence``.
            * ``encoding_low_confidence`` -- ``True`` when the detected encoding
              is inherently ambiguous (a single-byte/legacy codepage), the
              confidence is marginal, or several close alternatives exist. When
              set, a byte-exact read requires the caller to pin ``encoding=``.
              ``False`` for a caller-pinned or binary (XLSX) read.
            * ``encoding_alternatives`` -- other plausible encodings from the
              detector's ranked results (best first), empty when none / not
              sniffed. These are the candidates an explicit ``encoding=`` may
              need to disambiguate among.
            * ``text_columns`` -- sorted list of columns forced to text.
            * ``rows_read`` -- number of data rows in ``df``.
            * ``source`` -- ``"csv"`` or ``"xlsx"``.
            * ``parquet_cache`` -- path the cleaned frame was written to, or
              ``None``.
            * ``anomalies`` -- list of human-readable anomaly notes: structural
              flags (a row count at the spreadsheet truncation ceiling, an
              all-null column) AND data-fidelity warnings (an XLSX numeric ID
              past ``2**53`` whose source Excel already lost precision on).
    """

    df: pd.DataFrame
    report: dict[str, Any] = field(default_factory=dict)


# Google-Sheets / common spreadsheet-export row ceiling (2**20 - 1). A load that
# lands exactly here is very likely truncated upstream; flag it (verified
# arithmetic in the research gate: 2**20 - 1 == 1048575).
_TRUNCATION_CEILING = 2**20 - 1


def _normalize_name(name: object) -> str:
    return str(name).strip().lower()


def _tokenize_name(column_name: object) -> list[str]:
    """Split a column name into lowercased alphanumeric tokens.

    Boundaries are non-alphanumeric separators AND camelCase / acronym /
    letter-digit humps, so ``"Account_ID"`` -> ``["account", "id"]``,
    ``"caseNumber"`` -> ``["case", "number"]``, ``"zip_code"`` ->
    ``["zip", "code"]``. Used by :func:`_is_id_like` to match short ambiguous
    patterns against WHOLE tokens rather than raw substrings.
    """
    text = str(column_name).strip()
    text = _CAMEL_BOUNDARY_RE.sub(" ", text)
    text = text.lower()
    return [tok for tok in _NON_ALNUM_RE.split(text) if tok]


def _is_id_like(column_name: object) -> bool:
    """True if a column NAME is ID-like by WHOLE-TOKEN matching.

    EVERY pattern (short or long) matches only when it equals a whole token, so
    none fire on a word that merely CONTAINS the pattern: ``valid``/``raids``/
    ``casein``/``plateau`` (short patterns) and ``microphone_level``/
    ``zipper_count``/``accountability_score`` (long patterns) all stay numeric.
    Genuine ID names still match via their tokens: ``zip``, ``zip_code``,
    ``zipcode``, ``phone_number``, ``account_id``, ``fips``, ``case_number``,
    ``caseNumber``, ``ssn``, ``plate``, ``dob``. An operator can force an
    unmatched edge-case ID column to text via ``text_columns``; every coerced
    column is listed in ``report["text_columns"]`` so a missed/extra one shows.
    """
    return any(tok in ID_LIKE_TOKEN_PATTERNS for tok in _tokenize_name(column_name))


def _resolve_text_columns(
    columns: Iterable[object],
    explicit: Iterable[str] | None,
) -> list[str]:
    """Columns to force to text: ID-like names plus caller-named ones.

    Matching is by normalized (stripped, lowercased) name so a caller can pass
    ``"Zip"`` and have it match a ``" zip "`` header. Returns the ORIGINAL
    column labels (as they appear in the frame), sorted for a stable report.
    """
    explicit_norm = {_normalize_name(c) for c in (explicit or ())}
    chosen = [
        col
        for col in columns
        if _is_id_like(col) or _normalize_name(col) in explicit_norm
    ]
    return sorted(chosen, key=lambda c: str(c))


# Encodings whose detection is NOT inherently ambiguous: a clean UTF read is
# self-validating (invalid byte sequences raise), and ASCII is a strict subset.
# Anything OUTSIDE this set is a single-byte / legacy codepage, where statistical
# detection routinely cannot tell e.g. cp775 from latin-1 from cp1250 -- those
# decode the SAME bytes into DIFFERENT characters without raising. Detecting one
# of those is flagged low-confidence so the caller can require an explicit
# encoding= for a byte-exact read.
_UNAMBIGUOUS_ENCODING_PREFIXES: tuple[str, ...] = ("utf", "ascii", "us-ascii")

# Below this 1 - chaos margin a detection is treated as low-confidence even for
# an otherwise-unambiguous family.
_CONFIDENCE_MARGIN = 0.85


def _is_ambiguous_encoding(encoding: str) -> bool:
    """True for single-byte / legacy codepages (the inherently ambiguous class).

    UTF-* and ASCII are self-validating and treated as unambiguous; everything
    else (cp775, latin-1/iso-8859-*, cp125x, mac_*, koi8-*, ...) is a single-byte
    codepage that statistical detection cannot pick reliably.
    """
    norm = encoding.strip().lower().replace("_", "-")
    return not any(norm.startswith(p) for p in _UNAMBIGUOUS_ENCODING_PREFIXES)


@dataclass
class _Detection:
    encoding: str
    confidence: float | None
    low_confidence: bool
    alternatives: list[str]


def _detect_encoding(path: Path, sample_size: int = 65_536) -> _Detection:
    """Sniff a file's encoding from a byte sample, HONESTLY flagging ambiguity.

    Reads up to ``sample_size`` bytes and runs ``charset-normalizer``. Returns a
    :class:`_Detection` with the best-effort ``encoding``, a rough
    ``confidence`` (``1 - chaos`` of the best match, ``None`` on fallback), the
    other plausible ``alternatives`` from the ranked results, and
    ``low_confidence`` -- set whenever the pick is a single-byte/legacy codepage
    (the inherently ambiguous class), the confidence is marginal, or several
    close alternatives exist. ``confidence`` does NOT discriminate among
    single-byte alternatives (cp775 and latin-1 can both score ``1.0`` on the
    same bytes); read it together with ``low_confidence``. Falls back to
    ``utf-8`` when the sample is empty or the detector returns nothing.
    """
    raw = path.read_bytes()[:sample_size]
    if not raw:
        return _Detection("utf-8", None, False, [])
    # Imported lazily so importing this module does not hard-require the
    # detector unless an auto-detect load actually runs.
    from charset_normalizer import from_bytes

    results = from_bytes(raw)
    best = results.best()
    if best is None or not best.encoding:
        return _Detection("utf-8", None, False, [])

    # charset-normalizer exposes "chaos" (lower is better); turn it into a
    # rough confidence in [0, 1] for the report.
    confidence: float | None = None
    chaos = getattr(best, "chaos", None)
    if isinstance(chaos, (int, float)):
        confidence = max(0.0, min(1.0, 1.0 - float(chaos)))

    # Other plausible encodings from the ranked results (best first), de-duped
    # and excluding the winner. These are the candidates an explicit encoding=
    # might need to disambiguate among.
    alternatives: list[str] = []
    for match in results:
        enc = getattr(match, "encoding", None)
        if enc and enc != best.encoding and enc not in alternatives:
            alternatives.append(enc)

    ambiguous = _is_ambiguous_encoding(best.encoding)
    marginal = confidence is not None and confidence < _CONFIDENCE_MARGIN
    has_close_alternatives = len(alternatives) > 0 and ambiguous
    low_confidence = ambiguous or marginal or has_close_alternatives

    return _Detection(best.encoding, confidence, low_confidence, alternatives)


# IEEE-754 doubles represent every integer exactly only up to 2**53; beyond it
# consecutive integers collapse onto the same float. An ID stored as a NUMBER in
# Excel is already a rounded double by the time openpyxl yields it, so a 17-digit
# ID arrives WRONG (12345678901234567 -> 12345678901234570) and no amount of
# stringifying recovers the original. Flag any whitelisted numeric value at or
# above this magnitude.
_SAFE_INT_PRECISION = 2**53


def _coerce_text_columns(df: pd.DataFrame, text_columns: list[str]) -> list[str]:
    """In-place: cast the given columns to leading-zero-safe strings.

    Used on the XLSX path (and as a belt-and-suspenders pass on CSV). A value
    that pandas already read as a float ``7054.0`` is rendered back to the
    integer-looking ``"7054"`` (not ``"7054.0"``); genuine missing cells stay
    missing rather than becoming the literal string ``"nan"``.

    Returns a list of human-readable precision WARNINGS: when a whitelisted
    column arrives as a number whose magnitude exceeds safe integer precision
    (``abs >= 2**53``), Excel already stored it as a rounded double, so the
    stringified value may not be the true ID. The best-available value is still
    stringified -- but never presented as exact silently.
    """
    warnings: list[str] = []
    for col in text_columns:
        if col not in df.columns:
            continue
        series = df[col]
        na_mask = series.isna()

        # Precision audit: a numeric cell at/above 2**53 has already lost exact
        # integer precision upstream (Excel stored it as a double). Detect on the
        # numeric value regardless of whether the dtype is float or object-int.
        def _exceeds_precision(v: object) -> bool:
            if isinstance(v, bool) or v is None:
                return False
            if isinstance(v, (int, float)):
                try:
                    return abs(float(v)) >= _SAFE_INT_PRECISION
                except (OverflowError, ValueError):
                    # An int too large to be a finite float overflowed a double
                    # long ago; treat as a precision concern.
                    return True
            return False

        if series.map(_exceeds_precision).any():
            warnings.append(
                f"column {col!r}: numeric ID exceeds 2^53 -- source Excel "
                f"stored it as a number and may have lost precision; obtain a "
                f"text-typed export"
            )

        # Numeric whole-number floats -> int-looking strings (drop the ".0").
        if pd.api.types.is_float_dtype(series):
            as_str = series.map(
                lambda v: "" if pd.isna(v) else (
                    str(int(v)) if float(v).is_integer() else str(v)
                )
            )
        else:
            as_str = series.astype("string").astype(object)
        as_str = as_str.where(~na_mask, other=pd.NA)
        df[col] = as_str.astype(object)
        # Restore NA sentinels as actual NaN/None for consistency.
        df.loc[na_mask, col] = None
    return warnings


def _apply_empty_null(df: pd.DataFrame) -> None:
    """In-place: turn empty / whitespace-only string cells into ``None``."""
    for col in df.columns:
        series = df[col]
        if series.dtype != object and not pd.api.types.is_string_dtype(series):
            continue
        stripped_is_empty = series.map(
            lambda v: isinstance(v, str) and v.strip() == ""
        )
        if stripped_is_empty.any():
            df.loc[stripped_is_empty, col] = None


def _detect_anomalies(df: pd.DataFrame) -> list[str]:
    """Cheap, non-fatal anomaly notes for the report."""
    anomalies: list[str] = []
    n = len(df)
    if n == _TRUNCATION_CEILING:
        anomalies.append(
            f"row count is exactly the spreadsheet export ceiling "
            f"({_TRUNCATION_CEILING}); data is likely truncated upstream -- "
            f"request the gap"
        )
    # Columns that are entirely null after cleaning are worth surfacing.
    for col in df.columns:
        if len(df) and df[col].isna().all():
            anomalies.append(f"column {col!r} is entirely null")
    return anomalies


def _load_csv(
    path: Path,
    *,
    encoding: str | None,
    text_columns: list[str] | None,
    empty_null: bool,
    skiprows: int | list[int] | None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """CSV path: encoding pre-flight, then a single typed ``read_csv``.

    The text-whitelist is resolved in two passes because we cannot know the
    column names until we have read the header. First read the header row to
    learn the columns (cheap), resolve which are ID-like / caller-named, then
    read for real with ``dtype=str`` on exactly those columns.
    """
    detected = False
    confidence: float | None = None
    low_confidence = False
    alternatives: list[str] = []
    used_encoding = encoding
    if used_encoding is None:
        det = _detect_encoding(path)
        used_encoding = det.encoding
        confidence = det.confidence
        low_confidence = det.low_confidence
        alternatives = det.alternatives
        detected = True

    # Pass 1: read just the header (nrows=0) to learn the column names, honoring
    # skiprows so junk preamble does not masquerade as the header.
    header_df = pd.read_csv(
        path,
        encoding=used_encoding,
        skiprows=skiprows,
        nrows=0,
    )
    resolved_text = _resolve_text_columns(header_df.columns, text_columns)

    # Pass 2: the real read. dtype=str on the whitelisted columns preserves
    # leading zeros. empty_null is implemented with a NARROW NA set: ONLY a
    # literal empty / whitespace cell becomes NA. keep_default_na is FALSE so
    # pandas does NOT additionally apply its default NA spellings -- a cell that
    # literally contains "N/A", "NULL", "NaN", "None" survives as that STRING,
    # never silently nulled before derive_has_case / keyword matching sees it.
    # (Sweeping the default spellings would contradict the module's narrow-NA
    # contract; see _EMPTY_NA_VALUES.) With empty_null=False we keep "" literally
    # in str columns (post-fixed below via na_filter=False).
    dtype = {col: str for col in resolved_text}
    read_kwargs: dict[str, Any] = {
        "encoding": used_encoding,
        "skiprows": skiprows,
        "dtype": dtype,
    }
    if empty_null:
        # ONLY empty/whitespace cells are NA; default NA spellings are kept as
        # their literal strings (keep_default_na=False is load-bearing here).
        read_kwargs["na_values"] = list(_EMPTY_NA_VALUES)
        read_kwargs["keep_default_na"] = False
    else:
        # Opt out of ""->NA as well: disable all of pandas' default NA tokens and
        # add nothing to na_values. We restore literal "" in str columns below.
        read_kwargs["keep_default_na"] = False
        read_kwargs["na_values"] = []

    df = pd.read_csv(path, **read_kwargs)

    if not empty_null:
        # Restore literal "" in the whitelisted text columns (read_csv may have
        # NA'd them). Non-text columns are left as pandas read them.
        df = _restore_empty_strings_for_text(path, df, resolved_text, used_encoding, skiprows)

    report = {
        "source": "csv",
        "encoding": used_encoding,
        "encoding_used": used_encoding,
        "encoding_detected": detected,
        "encoding_confidence": confidence,
        "encoding_low_confidence": low_confidence,
        "encoding_alternatives": alternatives,
        "text_columns": resolved_text,
        # CSV reads ID columns as text bytes (dtype=str), so no double-rounding
        # precision loss is possible on this path; kept for report-shape parity.
        "_precision_warnings": [],
    }
    return df, report


def _restore_empty_strings_for_text(
    path: Path,
    df: pd.DataFrame,
    text_columns: list[str],
    encoding: str,
    skiprows: int | list[int] | None,
) -> pd.DataFrame:
    """For ``empty_null=False``: put literal ``""`` back in text columns.

    ``read_csv`` collapses ``""`` to ``NaN`` for object columns even with
    ``na_values=[]``. To honor ``empty_null=False`` we re-read with the C
    parser's ``na_filter=False`` (which disables ALL NA conversion) and copy the
    text columns across, leaving the numeric columns from the primary read
    untouched.
    """
    raw = pd.read_csv(
        path,
        encoding=encoding,
        skiprows=skiprows,
        dtype={col: str for col in text_columns},
        na_filter=False,
    )
    for col in text_columns:
        if col in raw.columns and col in df.columns:
            df[col] = raw[col].astype(object)
    return df


def _unmerge_and_fill(path: Path):
    """Load an XLSX with merged regions unmerged and filled from the top-left.

    Returns an in-memory workbook bytes buffer ready for ``pd.read_excel``.
    openpyxl drops every non-top-left cell value when a range is merged; this
    reads the top-left, unmerges, and writes that value into every cell of the
    former range -- handling vertical and horizontal merges and merged headers.
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(path)
    for ws in wb.worksheets:
        # Copy to a list first: unmerge_cells mutates the live ranges set.
        for mr in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = mr.bounds  # (col, row) order
            top_left = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(mr))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _load_xlsx(
    path: Path,
    *,
    text_columns: list[str] | None,
    empty_null: bool,
    skiprows: int | list[int] | None,
    sheet_name: int | str,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """XLSX path: unmerge+fill, read with openpyxl, then post-read coercion.

    The text-whitelist is applied as a post-read coercion (pandas/openpyxl would
    int-ify a leading-zero column), and ``empty_null`` as a post-read sweep. The
    deterministic unmerge pass (rather than a bare ``df.ffill()``) is preferred
    because it also fixes merged *headers* and horizontal merges; a plain ffill
    would only handle the "category merged down one column" case.
    """
    buf = _unmerge_and_fill(path)
    df = pd.read_excel(
        buf,
        sheet_name=sheet_name,
        engine="openpyxl",
        skiprows=skiprows,
        dtype=object,  # read everything raw; we coerce the whitelist ourselves
    )

    resolved_text = _resolve_text_columns(df.columns, text_columns)
    precision_warnings = _coerce_text_columns(df, resolved_text)

    if empty_null:
        _apply_empty_null(df)

    report = {
        "source": "xlsx",
        "encoding": "binary",  # XLSX is a zip container; no text encoding sniff
        "encoding_used": "binary",
        "encoding_detected": False,
        "encoding_confidence": None,
        # A binary read has no text-encoding ambiguity; flag it as confident.
        "encoding_low_confidence": False,
        "encoding_alternatives": [],
        "text_columns": resolved_text,
        # Surfaced into report["anomalies"] by load_table (see _detect_anomalies).
        "_precision_warnings": precision_warnings,
    }
    return df, report


def load_table(
    path,
    *,
    encoding: str | None = None,
    text_columns: list[str] | None = None,
    empty_null: bool = True,
    skiprows: int | list[int] | None = None,
    sheet_name: int | str = 0,
    forward_fill_columns: list[str] | None = None,
    parquet_cache=None,
) -> LoadResult:
    """Load a dirty CSV/XLSX into a clean DataFrame plus a load report.

    Args:
        path: path to a ``.csv`` / ``.tsv`` / ``.xlsx`` / ``.xls`` file. The
            suffix selects the path (CSV vs XLSX); unknown suffixes are treated
            as CSV.
        encoding: when ``None`` (default), the CSV encoding is sniffed from a
            byte sample and recorded. Single-byte/legacy codepages are
            inherently ambiguous, so a sniffed result is flagged in the report
            (``encoding_low_confidence``, ``encoding_alternatives``) rather than
            trusted silently; pass e.g. ``"latin-1"`` for a byte-exact read.
            Ignored for XLSX (a binary container).
        text_columns: extra column names to force to text beyond the ID-like
            auto-whitelist (matched case-insensitively by normalized name).
        empty_null: when ``True`` (default), empty-string cells become missing
            (``NaN``) consistently; ``False`` keeps literal ``""`` in text
            columns.
        skiprows: int or list of 0-based row indices to drop before the header
            (junk preamble rows).
        sheet_name: XLSX sheet index or name (default first sheet).
        forward_fill_columns: column names to forward-fill (``df.ffill()``)
            after cleaning -- the quick path for a "category label runs down a
            column, blank means same-as-above" layout, e.g. a CSV exported from
            a report that left the repeated label blank. Applied AFTER
            ``empty_null`` so blanks-turned-NaN are carried down. (The XLSX path
            already unmerges+fills merged regions deterministically; this is for
            the cases that arrive as plain blanks rather than true merges.)
        parquet_cache: when given, the cleaned frame is written to this path as
            Parquet and the path is noted in the report.

    Returns:
        A :class:`LoadResult` with ``.df`` and ``.report``.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm", ".xls"):
        df, report = _load_xlsx(
            path,
            text_columns=text_columns,
            empty_null=empty_null,
            skiprows=skiprows,
            sheet_name=sheet_name,
        )
    else:
        df, report = _load_csv(
            path,
            encoding=encoding,
            text_columns=text_columns,
            empty_null=empty_null,
            skiprows=skiprows,
        )

    df = df.reset_index(drop=True)

    # Forward-fill named label columns (the quick path for blank-means-repeat
    # layouts). pandas 3.0.3 removed fillna(method="ffill") -- it raises
    # TypeError -- so use the dedicated df[col].ffill(). Runs after empty_null so
    # "" has already become NaN and gets carried down.
    filled: list[str] = []
    for col in forward_fill_columns or ():
        if col in df.columns:
            df[col] = df[col].ffill()
            filled.append(col)
    report["forward_filled_columns"] = sorted(filled, key=str)

    report["rows_read"] = len(df)
    # Precision warnings (e.g. an Excel big-int ID past 2**53) are gathered at
    # read time, before reset_index; surface them alongside the cheap structural
    # anomaly notes so a published ID is never presented as exact silently.
    precision_warnings = report.pop("_precision_warnings", [])
    report["anomalies"] = list(precision_warnings) + _detect_anomalies(df)
    report["parquet_cache"] = None

    if parquet_cache is not None:
        cache_path = Path(parquet_cache)
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        # pandas -> pyarrow engine; preserves the text dtype of ID columns.
        df.to_parquet(cache_path, engine="pyarrow", index=False)
        report["parquet_cache"] = str(cache_path)

    return LoadResult(df=df, report=report)
